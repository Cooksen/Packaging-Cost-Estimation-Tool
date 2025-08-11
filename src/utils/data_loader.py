"""
data_loader.py

This module provides utility functions for loading and processing data files used in cost estimation workflows.

Functions included:
- `load_json(path)`: Load and parse a JSON file.
- `process_excel_to_json(file_path)`: Extract component pricing data (corrugate, EPE, MPP) from an Excel file and save them as structured JSON files.
- `process_freight_data(file_path)`: Extract freight cost data from an Excel file and save it as a JSON file.

All output files are saved under the `parsed_data/` directory.
"""

import json
import openpyxl

def load_json(path):
    """
    Load a JSON file from the specified path.

    Args:
        path (str): The path to the JSON file.

    Returns:
        dict: The loaded JSON data.
    """
    with open(path, "r") as file:
        data = json.load(file)
    return data

def process_excel_to_json(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    corrugate_data, epe_data, mpp_data, bag_data = [], [], [], []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        material = row_dict.get("material", "")
        if material is not None:
            material = material.strip().lower()
            if material == "corrugate":
                corrugate_data.append(row_dict)
            elif material == "epe":
                epe_data.append(row_dict)
            elif material == "mpp":
                mpp_data.append(row_dict)
            elif material == "bag":
                bag_data.append(row_dict)

    # Dump JSON files
    with open("parsed_data/price_weight_corrugate.json", "w", encoding="utf-8") as f:
        json.dump(corrugate_data, f, ensure_ascii=False, indent=2)
    with open("parsed_data/price_weight_EPE.json", "w", encoding="utf-8") as f:
        json.dump(epe_data, f, ensure_ascii=False, indent=2)
    with open("parsed_data/price_weight_MPP.json", "w", encoding="utf-8") as f:
        json.dump(mpp_data, f, ensure_ascii=False, indent=2)
    with open("parsed_data/price_size_bag.json", "w", encoding="utf-8") as f:
        json.dump(bag_data, f, ensure_ascii=False, indent=2)

    return {
        "Corrugate": len(corrugate_data),
        "EPE": len(epe_data),
        "MPP": len(mpp_data),
        "Bag": len(bag_data)
    }

def process_freight_data(file_path):
    """
    Process freight cost data from an Excel file and save it as a JSON file.

    Args:
        file_path (str): Path to the uploaded Excel file.
        output_path (str): Destination path for the JSON file.

    Returns:
        int: Number of records processed.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    freight_data = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            row_dict = dict(zip(headers, row))
            freight_data.append(row_dict)

    output_path = "parsed_data/freight_cost_data.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(freight_data, f, ensure_ascii=False, indent=2)

    return len(freight_data)